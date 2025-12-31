from __future__ import annotations

from datetime import datetime
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.specialty import Specialty
from app.repo.specialties import SpecialtyRepo
from app.schemas.category import CategoryBatchAction, CategoryBatchResult, CategoryCreate, CategoryUpdate
from app.schemas.pagination import PageParams
from app.services import specialties as specialty_service

EXPORT_HEADERS = ["专业编码", "专业名称", "上级编码", "启用", "排序"]
HEADER_MAP = {
    "专业编码": "code",
    "编码": "code",
    "code": "code",
    "专业名称": "name",
    "名称": "name",
    "name": "name",
    "上级编码": "parent_code",
    "parent_code": "parent_code",
    "parent code": "parent_code",
    "启用": "is_active",
    "active": "is_active",
    "排序": "sort_order",
    "sort": "sort_order",
    "sort_order": "sort_order",
}


def _coerce_str(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _coerce_bool(value: object | None, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return bool(value)
    if isinstance(value, float):
        return bool(int(value))
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "y", "是"}:
            return True
        if normalized in {"0", "false", "no", "n", "否"}:
            return False
    return default


def _coerce_int(value: object | None, default: int = 0) -> int:
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return default


def list_categories(db: Session, params: PageParams) -> tuple[list[Specialty], int]:
    return specialty_service.list_specialties(db, params)


def list_category_tree(db: Session):
    return specialty_service.list_specialty_tree(db)


def create_category(db: Session, payload: CategoryCreate) -> Specialty:
    return specialty_service.create_specialty(db, payload)


def update_category(db: Session, category_id: int, payload: CategoryUpdate) -> Specialty:
    return specialty_service.update_specialty(db, category_id, payload)


def delete_category(db: Session, category_id: int) -> None:
    specialty_service.delete_specialty(db, category_id)


def batch_categories(db: Session, payload: CategoryBatchAction) -> CategoryBatchResult:
    ids = [item.id for item in payload.items]
    result = specialty_service.batch_specialties(db, payload.action, ids)
    return CategoryBatchResult(
        updated=int(result.get("updated", 0)),
        deleted=int(result.get("deleted", 0)),
        skipped=int(result.get("skipped", 0)),
        errors=result.get("errors", []),
    )


def import_categories(db: Session, file) -> dict[str, int | list[dict[str, object]]]:
    file.file.seek(0)
    workbook = load_workbook(file.file, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    index_to_field: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip()
        field = HEADER_MAP.get(key) or HEADER_MAP.get(key.lower())
        if field:
            index_to_field[idx] = field

    if not index_to_field:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Missing valid headers",
        )

    entries: list[dict[str, object]] = []
    for row_index, row in enumerate(rows, start=2):
        if not row or all(cell is None for cell in row):
            continue
        data: dict[str, object | None] = {"row": row_index}
        for idx, field in index_to_field.items():
            value = row[idx] if idx < len(row) else None
            data[field] = value
        code = _coerce_str(data.get("code"))
        name = _coerce_str(data.get("name"))
        if not code or not name:
            entries.append({"row": row_index, "detail": "Missing code or name"})
            continue
        entries.append(
            {
                "row": row_index,
                "code": code,
                "name": name,
                "parent_code": _coerce_str(data.get("parent_code")),
                "is_active": _coerce_bool(data.get("is_active"), True),
                "sort_order": _coerce_int(data.get("sort_order"), 0),
            }
        )

    if not entries:
        return {"created": 0, "updated": 0, "skipped": 0, "errors": []}

    repo = SpecialtyRepo(db)
    by_code = {item.code: item for item in repo.list() if item.code}
    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, object]] = []
    pending = entries[:]
    progress = True

    while pending and progress:
        progress = False
        for entry in pending[:]:
            parent_code = entry.get("parent_code")
            parent_id = None
            if parent_code:
                parent = by_code.get(parent_code)
                if parent is None:
                    continue
                parent_id = parent.id
            if parent_code and parent_id is None:
                continue

            existing = by_code.get(entry["code"])
            payload = {
                "parent_id": parent_id,
                "name": entry["name"],
                "code": entry["code"],
                "is_active": entry["is_active"],
                "sort_order": entry["sort_order"],
            }
            if existing:
                specialty_service.update_specialty(db, existing.id, CategoryUpdate(**payload))
                updated += 1
                by_code[entry["code"]] = existing
            else:
                created_item = specialty_service.create_specialty(
                    db, CategoryCreate(**payload)
                )
                by_code[entry["code"]] = created_item
                created += 1
            pending.remove(entry)
            progress = True

    for entry in pending:
        skipped += 1
        errors.append(
            {
                "row": entry.get("row"),
                "detail": "Parent code not found",
            }
        )

    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


def export_categories(db: Session) -> BytesIO:
    specialties = SpecialtyRepo(db).list()
    by_id = {item.id: item for item in specialties}
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(EXPORT_HEADERS)

    for item in sorted(specialties, key=lambda entry: (entry.sort_order, entry.id)):
        parent_code = None
        if item.parent_id and item.parent_id in by_id:
            parent_code = by_id[item.parent_id].code
        worksheet.append(
            [
                item.code or "",
                item.name,
                parent_code or "",
                "是" if item.is_active else "否",
                item.sort_order,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


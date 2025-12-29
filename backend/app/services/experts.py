from __future__ import annotations

from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.expert import Expert
from app.models.expert_document import ExpertDocument
from app.models.expert_specialty import ExpertSpecialty
from app.models.specialty import Specialty
from app.repo.experts import ExpertRepo
from app.repo.specialties import SpecialtyRepo
from app.schemas.pagination import PageParams
from app.services import organizations as organization_service
from app.services import titles as title_service
from app.services import categories as category_service
from app.services import specialties as specialty_service
from app.services import regions as region_service
from app.schemas.expert import ExpertCreate, ExpertUpdate

APPOINTMENT_DOC_TYPE = "appointment_letter"

EXPORT_FIELDS = [
    ("name", "姓名"),
    ("gender", "性别"),
    ("id_card_no", "身份证号"),
    ("phone", "电话"),
    ("email", "邮箱"),
    ("company", "单位"),
    ("region", "地域"),
    ("title", "职称"),
    ("specialty_codes", "专业编码"),
    ("appointment_letter_urls", "聘书图片"),
    ("is_active", "启用"),
]
EXPORT_HEADERS = [label for _, label in EXPORT_FIELDS]
HEADER_MAP = {label.lower(): field for field, label in EXPORT_FIELDS}
LEGACY_HEADER_MAP = {
    "name": "name",
    "gender": "gender",
    "idcardno": "id_card_no",
    "id_card_no": "id_card_no",
    "id card": "id_card_no",
    "phone": "phone",
    "email": "email",
    "company": "company",
    "region": "region",
    "title": "title",
    "category": "category",
    "subcategory": "subcategory",
    "specialty": "specialties",
    "specialties": "specialties",
    "specialty_codes": "specialty_codes",
    "specialty codes": "specialty_codes",
    "appointment_letter_urls": "appointment_letter_urls",
    "appointment letters": "appointment_letter_urls",
    "isactive": "is_active",
    "is active": "is_active",
}
HEADER_MAP.update(LEGACY_HEADER_MAP)
HEADER_MAP.update({field: field for field, _ in EXPORT_FIELDS})


def _coerce_str(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _coerce_bool(value: object | None, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return bool(value)
    if isinstance(value, float):
        return bool(int(value))
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "y"}:
            return True
        if normalized in {"0", "false", "no", "n"}:
            return False
    return default


def _split_list(value: object | None) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        normalized = value.replace("、", ";").replace("|", ";")
        parts = [item.strip() for item in normalized.split(";")]
        return [item for item in parts if item]
    if isinstance(value, (int, float)):
        return [str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)]
    return [str(value).strip()] if str(value).strip() else []


def _mask_phone(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if len(raw) <= 7:
        return raw
    return f"{raw[:3]}{'*' * (len(raw) - 7)}{raw[-4:]}"


def _mask_id_card(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if len(raw) <= 7:
        return raw
    return f"{raw[:3]}{'*' * (len(raw) - 7)}{raw[-4:]}"


def _mask_name(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    if len(raw) == 1:
        return raw
    if len(raw) == 2:
        return f"{raw[0]}*"
    return f"{raw[0]}{'*' * (len(raw) - 2)}{raw[-1]}"


def _ensure_id_card_unique(
    db: Session, id_card_no: str | None, exclude_id: int | None = None
) -> None:
    if not id_card_no:
        return
    stmt = select(Expert.id).where(Expert.id_card_no == id_card_no)
    if exclude_id is not None:
        stmt = stmt.where(Expert.id != exclude_id)
    exists = db.execute(stmt).first() is not None
    if exists:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID card already exists",
        )


def _attach_expert_details(db: Session, experts: list[Expert]) -> None:
    if not experts:
        return
    expert_ids = [expert.id for expert in experts if expert.id is not None]
    if not expert_ids:
        return

    specialty_map: dict[int, list[Specialty]] = {expert_id: [] for expert_id in expert_ids}
    specialty_stmt = (
        select(ExpertSpecialty.expert_id, Specialty)
        .join(Specialty, Specialty.id == ExpertSpecialty.specialty_id)
        .where(ExpertSpecialty.expert_id.in_(expert_ids))
        .order_by(Specialty.sort_order, Specialty.id)
    )
    for expert_id, specialty in db.execute(specialty_stmt).all():
        specialty_map.setdefault(expert_id, []).append(specialty)

    doc_map: dict[int, list[str]] = {}
    doc_stmt = (
        select(ExpertDocument)
        .where(
            ExpertDocument.expert_id.in_(expert_ids),
            ExpertDocument.doc_type == APPOINTMENT_DOC_TYPE,
        )
        .order_by(ExpertDocument.sort_order, ExpertDocument.id)
    )
    for doc in db.execute(doc_stmt).scalars().all():
        doc_map.setdefault(doc.expert_id, []).append(doc.url)

    for expert in experts:
        specialties = specialty_map.get(expert.id, [])
        setattr(expert, "specialties", specialties)
        setattr(expert, "specialty_ids", [item.id for item in specialties])
        setattr(expert, "appointment_letter_urls", doc_map.get(expert.id, []))


def _sync_expert_specialties(
    db: Session, expert_id: int, specialty_ids: list[int] | None
) -> None:
    if specialty_ids is None:
        return
    unique_ids: list[int] = []
    for item in specialty_ids:
        try:
            value = int(item)
        except (TypeError, ValueError):
            continue
        if value not in unique_ids:
            unique_ids.append(value)

    if not unique_ids:
        db.execute(
            delete(ExpertSpecialty).where(ExpertSpecialty.expert_id == expert_id)
        )
        return

    existing = set(
        db.execute(select(Specialty.id).where(Specialty.id.in_(unique_ids)))
        .scalars()
        .all()
    )
    missing = [item for item in unique_ids if item not in existing]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Specialty not found"
        )

    current = set(
        db.execute(
            select(ExpertSpecialty.specialty_id).where(
                ExpertSpecialty.expert_id == expert_id
            )
        )
        .scalars()
        .all()
    )
    to_remove = [item for item in current if item not in unique_ids]
    if to_remove:
        db.execute(
            delete(ExpertSpecialty).where(
                ExpertSpecialty.expert_id == expert_id,
                ExpertSpecialty.specialty_id.in_(to_remove),
            )
        )
    to_add = [item for item in unique_ids if item not in current]
    for specialty_id in to_add:
        db.add(ExpertSpecialty(expert_id=expert_id, specialty_id=specialty_id))


def _sync_expert_documents(
    db: Session, expert_id: int, urls: list[str] | None
) -> None:
    if urls is None:
        return
    clean_urls = [item.strip() for item in urls if isinstance(item, str) and item.strip()]
    db.execute(
        delete(ExpertDocument).where(
            ExpertDocument.expert_id == expert_id,
            ExpertDocument.doc_type == APPOINTMENT_DOC_TYPE,
        )
    )
    for index, url in enumerate(clean_urls, start=1):
        db.add(
            ExpertDocument(
                expert_id=expert_id,
                doc_type=APPOINTMENT_DOC_TYPE,
                url=url,
                sort_order=index,
            )
        )


def list_experts(db: Session, params: PageParams) -> tuple[list[Expert], int]:
    items, total = ExpertRepo(db).list_page(
        params.keyword,
        params.sort_by,
        params.sort_order,
        params.page,
        params.page_size,
    )
    _attach_expert_details(db, items)
    return items, total


def list_experts_all(db: Session) -> list[Expert]:
    return ExpertRepo(db).list()


def get_expert(db: Session, expert_id: int) -> Expert:
    expert = ExpertRepo(db).get_by_id(expert_id)
    if expert is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Expert not found"
        )
    _attach_expert_details(db, [expert])
    return expert


def create_expert(db: Session, payload: ExpertCreate) -> Expert:
    data = payload.model_dump()
    specialty_ids = data.pop("specialty_ids", [])
    appointment_letter_urls = data.pop("appointment_letter_urls", [])

    if not data.get("id_card_no"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID card is required",
        )
    _ensure_id_card_unique(db, data.get("id_card_no"))
    organization = organization_service.resolve_organization(
        db,
        data.get("organization_id"),
        data.get("company"),
        strict=data.get("organization_id") is not None,
        create_if_missing=bool(data.get("company"))
        and data.get("organization_id") is None,
    )
    region = region_service.resolve_region(
        db,
        data.get("region_id"),
        data.get("region"),
        strict=data.get("region_id") is not None,
        create_if_missing=bool(data.get("region"))
        and data.get("region_id") is None,
    )
    title = title_service.resolve_title(
        db,
        data.get("title_id"),
        data.get("title"),
        strict=data.get("title_id") is not None,
    )

    expert = Expert(**data)
    if organization:
        expert.organization_id = organization.id
        expert.company = organization.name
    if region:
        expert.region_id = region.id
        expert.region = region.name
    if title:
        expert.title_id = title.id
        expert.title = title.name
    db.add(expert)
    db.flush()
    _sync_expert_specialties(db, expert.id, specialty_ids)
    _sync_expert_documents(db, expert.id, appointment_letter_urls)
    db.commit()
    db.refresh(expert)
    _attach_expert_details(db, [expert])
    return expert


def update_expert(db: Session, expert_id: int, payload: ExpertUpdate) -> Expert:
    expert = get_expert(db, expert_id)
    update_data = payload.model_dump(exclude_unset=True)
    organization_input = (
        "organization_id" in update_data or "company" in update_data
    )
    region_input = "region_id" in update_data or "region" in update_data
    title_input = "title_id" in update_data or "title" in update_data
    specialty_ids = update_data.pop("specialty_ids", None)
    appointment_letter_urls = update_data.pop("appointment_letter_urls", None)

    if "id_card_no" in update_data:
        if not update_data.get("id_card_no"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="ID card is required",
            )
        _ensure_id_card_unique(db, update_data.get("id_card_no"), exclude_id=expert_id)

    if organization_input:
        if update_data.get("organization_id") is None and update_data.get("company") is None:
            expert.organization_id = None
            expert.company = None
        else:
            organization = organization_service.resolve_organization(
                db,
                update_data.get("organization_id"),
                update_data.get("company"),
                strict=update_data.get("organization_id") is not None,
                create_if_missing=bool(update_data.get("company"))
                and update_data.get("organization_id") is None,
            )
            if organization:
                expert.organization_id = organization.id
                expert.company = organization.name
            else:
                expert.organization_id = update_data.get("organization_id")
                if "company" in update_data:
                    expert.company = update_data.get("company")

    if region_input:
        if update_data.get("region_id") is None and update_data.get("region") is None:
            expert.region_id = None
            expert.region = None
        else:
            region = region_service.resolve_region(
                db,
                update_data.get("region_id"),
                update_data.get("region"),
                strict=update_data.get("region_id") is not None,
                create_if_missing=bool(update_data.get("region"))
                and update_data.get("region_id") is None,
            )
            if region:
                expert.region_id = region.id
                expert.region = region.name
            else:
                expert.region_id = update_data.get("region_id")
                if "region" in update_data:
                    expert.region = update_data.get("region")

    if title_input:
        if update_data.get("title_id") is None and update_data.get("title") is None:
            expert.title_id = None
            expert.title = None
        else:
            title = title_service.resolve_title(
                db,
                update_data.get("title_id"),
                update_data.get("title"),
                strict=update_data.get("title_id") is not None,
            )
            if title:
                expert.title_id = title.id
                expert.title = title.name
            else:
                expert.title_id = update_data.get("title_id")
                if "title" in update_data:
                    expert.title = update_data.get("title")

    for key, value in update_data.items():
        if key in {"organization_id", "company", "region_id", "region", "title_id", "title"}:
            continue
        setattr(expert, key, value)
    _sync_expert_specialties(db, expert_id, specialty_ids)
    _sync_expert_documents(db, expert_id, appointment_letter_urls)
    db.commit()
    db.refresh(expert)
    _attach_expert_details(db, [expert])
    return expert


def delete_expert(db: Session, expert_id: int) -> None:
    expert = get_expert(db, expert_id)
    db.execute(
        delete(ExpertSpecialty).where(ExpertSpecialty.expert_id == expert_id)
    )
    db.execute(
        delete(ExpertDocument).where(ExpertDocument.expert_id == expert_id)
    )
    db.delete(expert)
    db.commit()


def delete_experts(db: Session, expert_ids: list[int]) -> dict[str, int]:
    unique_ids = {int(item) for item in expert_ids if isinstance(item, int)}
    if not unique_ids:
        return {"deleted": 0, "skipped": 0}
    existing = set(
        db.execute(select(Expert.id).where(Expert.id.in_(unique_ids)))
        .scalars()
        .all()
    )
    if not existing:
        return {"deleted": 0, "skipped": len(unique_ids)}

    db.execute(
        delete(ExpertSpecialty).where(ExpertSpecialty.expert_id.in_(existing))
    )
    db.execute(
        delete(ExpertDocument).where(ExpertDocument.expert_id.in_(existing))
    )
    db.execute(delete(Expert).where(Expert.id.in_(existing)))
    db.commit()
    return {"deleted": len(existing), "skipped": len(unique_ids) - len(existing)}


def import_experts(db: Session, file) -> dict[str, int]:
    file.file.seek(0)
    workbook = load_workbook(file.file, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return {"created": 0, "skipped": 0}

    index_to_field: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        field = HEADER_MAP.get(key)
        if field:
            index_to_field[idx] = field

    if not index_to_field:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Missing valid headers",
        )

    created = 0
    skipped = 0
    for row_index, row in enumerate(rows, start=2):
        if not row or all(cell is None for cell in row):
            continue
        data: dict[str, object | None] = {}
        for idx, field in index_to_field.items():
            value = row[idx] if idx < len(row) else None
            data[field] = value

        name = _coerce_str(data.get("name"))
        if not name:
            skipped += 1
            continue

        id_card_no = _coerce_str(data.get("id_card_no"))
        if not id_card_no:
            skipped += 1
            continue
        existing = db.execute(
            select(Expert.id).where(Expert.id_card_no == id_card_no)
        ).first()
        if existing:
            skipped += 1
            continue

        region_name = _coerce_str(data.get("region"))
        expert = Expert(
            name=name,
            id_card_no=id_card_no,
            gender=_coerce_str(data.get("gender")),
            phone=_coerce_str(data.get("phone")),
            email=_coerce_str(data.get("email")),
            company=_coerce_str(data.get("company")),
            region=region_name,
            title=_coerce_str(data.get("title")),
            is_active=_coerce_bool(data.get("is_active"), True),
        )
        category_name = _coerce_str(data.get("category"))
        subcategory_name = _coerce_str(data.get("subcategory"))
        category = category_service.resolve_category(
            db,
            None,
            category_name,
            strict=False,
            create_if_missing=False,
        )
        subcategory = category_service.resolve_subcategory(
            db,
            None,
            subcategory_name,
            category,
            strict=False,
            create_if_missing=False,
        )
        organization = organization_service.resolve_organization(
            db,
            None,
            expert.company,
            strict=False,
            create_if_missing=True,
        )
        region = region_service.resolve_region(
            db,
            None,
            region_name,
            strict=False,
            create_if_missing=bool(region_name),
        )
        title = title_service.resolve_title(
            db,
            None,
            expert.title,
            strict=False,
            create_if_missing=True,
        )
        if subcategory and category is None:
            category = category_service.resolve_category(
                db, subcategory.category_id, None, strict=False
            )
        if organization:
            expert.organization_id = organization.id
            expert.company = organization.name
        if region:
            expert.region_id = region.id
            expert.region = region.name
        if title:
            expert.title_id = title.id
            expert.title = title.name
        db.add(expert)
        db.flush()

        specialty_names = _split_list(data.get("specialties"))
        specialty_codes = _split_list(data.get("specialty_codes"))
        if specialty_names and not specialty_codes:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"请使用专业编码 (第{row_index}行)",
            )
        specialty_ids: list[int] = []
        if specialty_codes:
            missing_codes: list[str] = []
            for code in specialty_codes:
                specialty = SpecialtyRepo(db).get_by_code(code)
                if specialty is None:
                    missing_codes.append(code)
                    continue
                specialty_ids.append(specialty.id)
            if missing_codes:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"专业不存在: {', '.join(missing_codes)} (第{row_index}行)",
                )

        _sync_expert_specialties(db, expert.id, specialty_ids)
        appointment_letter_urls = _split_list(data.get("appointment_letter_urls"))
        _sync_expert_documents(db, expert.id, appointment_letter_urls)
        created += 1

    db.commit()
    return {"created": created, "skipped": skipped}


def export_experts(db: Session) -> BytesIO:
    experts = ExpertRepo(db).list()
    _attach_expert_details(db, experts)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(EXPORT_HEADERS)

    for expert in experts:
        specialties = getattr(expert, "specialties", [])
        specialty_names = ";".join([item.name for item in specialties if item.name])
        specialty_codes = ";".join([item.code for item in specialties if item.code])
        appointment_urls = ";".join(
            getattr(expert, "appointment_letter_urls", [])
        )
        row = []
        for field, _label in EXPORT_FIELDS:
            if field == "specialties":
                row.append(specialty_names)
            elif field == "specialty_codes":
                row.append(specialty_codes)
            elif field == "appointment_letter_urls":
                row.append(appointment_urls)
            elif field == "name":
                row.append(_mask_name(expert.name))
            elif field == "id_card_no":
                row.append(_mask_id_card(expert.id_card_no))
            elif field == "phone":
                row.append(_mask_phone(expert.phone))
            else:
                row.append(getattr(expert, field, None))
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
